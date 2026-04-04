#!/usr/bin/env python3
"""SIM 덤프 JSON을 Excel(.xlsx)로 변환."""

import json
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# 스타일
DEFAULT_FONT = Font(name="Consolas", size=11)
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(name="Consolas", size=11, color="0563C1", underline="single")
BACK_FONT = Font(name="Consolas", size=11, color="0563C1", underline="single", bold=True)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DF_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ERROR_FILL = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
WARN_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

MAX_PATH_DEPTH = 3
PATH_HEADERS = ["Level 1", "Level 2", "Level 3"]

STRUCTURE_MAP = {
    "transparent": "TF",
    "linear_fixed": "LF",
    "cyclic": "CF",
    "ber_tlv": "BER-TLV",
    "no_info_given": "",
}


def _sort_key(file_path: str, file_data: dict) -> str:
    """Sort files: MF → MF/EF.* → ADF.USIM(EFs first, then sub-DFs) → ADF.ISIM → DF.GSM → DF.TELECOM → rest.
       Within each group, sort by FID ascending."""
    if file_path == 'MF':
        group = '0'
    elif file_path.startswith('MF/EF.'):
        group = '0b'
    elif file_path.startswith('MF/ADF.USIM'):
        group = '1'
    elif file_path.startswith('MF/ADF.ISIM'):
        group = '2'
    elif file_path.startswith('MF/DF.GSM'):
        group = '3'
    elif file_path.startswith('MF/DF.TELECOM'):
        group = '4'
    elif file_path.startswith('MF/DF.'):
        group = '5'
    else:
        group = '6'

    parts = file_path.split('/')
    depth = len(parts)
    fid = (file_data.get('fcp', {}).get('file_identifier', '') or 'ffff').lower()

    sub = '0_' + fid
    if group in ('1', '2', '3', '4'):
        if depth == 2:
            sub = '0_0000'
        elif depth == 3 and not parts[2].startswith('DF.'):
            sub = '1_' + fid
        else:
            sub_df = parts[2] if len(parts) > 2 else ''
            if depth == 3:
                sub = '2_' + sub_df + '_0000'
            else:
                sub = '2_' + sub_df + '_' + fid

    return group + '_' + sub


def _sorted_files(data: dict) -> list[tuple[str, dict]]:
    """Return files sorted by the standard display order."""
    return sorted(data.get("files", {}).items(), key=lambda x: _sort_key(x[0], x[1]))


def _bytes_to_str(raw_bytes, max_len=500) -> str:
    if raw_bytes is None:
        return ""
    if isinstance(raw_bytes, str):
        return raw_bytes[:max_len].upper()
    if isinstance(raw_bytes, list):
        s = " | ".join(str(r).upper() for r in raw_bytes)
        return s[:max_len] + ("..." if len(s) > max_len else "")
    if isinstance(raw_bytes, dict):
        s = json.dumps(raw_bytes, ensure_ascii=False).upper()
        return s[:max_len]
    return str(raw_bytes)[:max_len].upper()


def _path_to_filename(path_parts: list) -> str:
    return "_".join(path_parts) + ".json"


def _make_sheet_name(path_parts: list) -> str:
    """시트 이름 생성 (31자 제한)."""
    # EF.IMSI 같은 마지막 이름 우선, 중복 방지를 위해 부모 포함
    name = path_parts[-1] if path_parts else "unknown"
    if len(path_parts) >= 3:
        name = f"{path_parts[-2]}_{name}"
    return name[:31]


def _save_decoded_files(data: dict, card_dir: Path) -> Path:
    decoded_dir = card_dir / "decoded"
    decoded_dir.mkdir(parents=True, exist_ok=True)
    for file_path, file_data in data.get("files", {}).items():
        body = file_data.get("body")
        if body is None:
            continue
        path_parts = file_data.get("path", [])
        filename = _path_to_filename(path_parts)
        with open(decoded_dir / filename, "w", encoding="utf-8") as f:
            json.dump(body, f, indent=2, ensure_ascii=False)
    return decoded_dir


def _get_file_info(file_data: dict) -> dict:
    fcp = file_data.get("fcp", {})
    fd = fcp.get("file_descriptor", {}).get("file_descriptor_byte", {})
    file_type = fd.get("file_type", "")
    error = file_data.get("error")

    if file_type == "df":
        type_str = "DF"
    elif file_type in ("working_ef", "internal_ef"):
        type_str = "EF"
    else:
        type_str = file_type

    structure = fd.get("structure", "")
    ef_type = STRUCTURE_MAP.get(structure, structure)

    path_parts = file_data.get("path", [])
    if path_parts == ["MF"]:
        display_parts = ["MF"]
    elif path_parts and path_parts[0] == "MF":
        display_parts = path_parts[1:]
    else:
        display_parts = path_parts
    path_cols = list(display_parts) + [""] * (MAX_PATH_DEPTH - len(display_parts))

    has_body = file_data.get("body") is not None
    json_filename = _path_to_filename(path_parts) if has_body else None

    # SW 에러 (6a82 등)와 디코딩 에러 분리
    sw_error = ""
    decode_error = ""
    if error:
        if "sw_actual" in error:
            sw_error = f"{error['sw_actual']}: {error.get('message', '')}"
        else:
            sw_error = f"ERR: {error.get('message', '')}"

    return {
        "path_cols": path_cols[:MAX_PATH_DEPTH],
        "fid": fcp.get("file_identifier", "").upper(),
        "type": type_str,
        "ef_type": ef_type,
        "file_size": fcp.get("file_size", ""),
        "record_len": fcp.get("file_descriptor", {}).get("record_len", ""),
        "num_of_rec": fcp.get("file_descriptor", {}).get("num_of_rec", ""),
        "raw_bytes": _bytes_to_str(file_data.get("bytes")),
        "json_filename": json_filename,
        "error": sw_error,
        "decode_error": decode_error,
    }


def _write_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _is_record_ef(file_data: dict) -> bool:
    """LF 또는 CF 파일인지 확인."""
    fd = file_data.get("fcp", {}).get("file_descriptor", {}).get("file_descriptor_byte", {})
    return fd.get("structure") in ("linear_fixed", "cyclic")


def _create_record_sheet(wb, file_data: dict, path_parts: list, files_row: int, used_names: set) -> str:
    """LF/CF 파일의 레코드를 별도 시트로 생성. 시트 이름 반환."""
    sheet_name = _make_sheet_name(path_parts)
    # 중복 방지
    base = sheet_name
    counter = 2
    while sheet_name in used_names:
        sheet_name = f"{base[:28]}_{counter}"
        counter += 1
    used_names.add(sheet_name)

    ws = wb.create_sheet(sheet_name)

    # 상단: Back 링크
    back_cell = ws.cell(row=1, column=1, value="← Back to Files")
    back_cell.hyperlink = f"#Files!A{files_row}"
    back_cell.font = BACK_FONT

    # 파일 경로 표시
    cell = ws.cell(row=2, column=1, value="/".join(path_parts))
    cell.font = DEFAULT_FONT

    # 헤더
    rec_headers = ["Record #", "Hex Data"]
    for c, h in enumerate(rec_headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # 레코드 데이터
    raw_records = file_data.get("bytes", [])
    if isinstance(raw_records, list):
        for i, rec in enumerate(raw_records, 1):
            c1 = ws.cell(row=4 + i, column=1, value=i)
            c1.border = BORDER
            c1.font = DEFAULT_FONT
            rec_str = str(rec).upper() if rec else ""
            c2 = ws.cell(row=4 + i, column=2, value=rec_str)
            c2.border = BORDER
            c2.font = DEFAULT_FONT

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 80

    return sheet_name


def convert_to_excel(json_path: str, output_path: str = None) -> str:
    json_path = Path(json_path)
    card_dir = json_path.parent
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if output_path is None:
        output_path = card_dir / "dump.xlsx"

    _save_decoded_files(data, card_dir)

    wb = Workbook()
    wb._named_styles['Normal'].font = DEFAULT_FONT

    # === Sheet 1: Card Info ===
    ws_info = wb.active
    ws_info.title = "Card Info"
    info_rows = [
        ("항목", "값"),
        ("Card Name", data.get("name", "")),
        ("ATR", data.get("atr", "")),
        ("ICCID", data.get("iccid", "")),
        ("EID", data.get("eid", "") or ""),
        ("Total Files", str(len(data.get("files", {})))),
        ("AIDs", ", ".join(data.get("aids", {}).keys())),
    ]
    for r, (key, val) in enumerate(info_rows, 1):
        ws_info.cell(row=r, column=1, value=key).font = DEFAULT_FONT
        ws_info.cell(row=r, column=2, value=val).font = DEFAULT_FONT
        if r == 1:
            for c in (1, 2):
                ws_info.cell(row=r, column=c).font = HEADER_FONT
                ws_info.cell(row=r, column=c).fill = HEADER_FILL
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 80

    # === Sheet 2: Files ===
    ws_files = wb.create_sheet("Files")
    headers = PATH_HEADERS + ["FID", "Type", "EF Type", "Size",
                               "Record Len", "Num Records",
                               "Bytes", "Decoding", "Error"]
    _write_header(ws_files, headers)
    bytes_col = len(headers) - 2   # Bytes 컬럼
    decoding_col = len(headers) - 1  # Decoding 컬럼

    # LF/CF 시트 생성을 위해 먼저 수집
    record_sheets = {}  # file_path -> (files_row, file_data)
    used_sheet_names = set()

    row = 2
    for file_path, file_data in _sorted_files(data):
        info = _get_file_info(file_data)
        values = (info["path_cols"] +
                  [info["fid"], info["type"], info["ef_type"],
                   info["file_size"], info["record_len"], info["num_of_rec"],
                   info["raw_bytes"], "", info["error"]])
        for c, val in enumerate(values, 1):
            cell = ws_files.cell(row=row, column=c)
            if isinstance(val, (int, float)) and val != "":
                cell.value = val
            else:
                cell.value = str(val) if val else ""
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top")
            cell.font = DEFAULT_FONT

        # LF/CF: Bytes 셀을 시트 링크로
        if _is_record_ef(file_data) and file_data.get("bytes"):
            record_sheets[file_path] = (row, file_data)

        # Decoding 열: JSON 링크 또는 디코딩 에러
        if info["json_filename"]:
            link_path = f"decoded/{info['json_filename']}"
            cell = ws_files.cell(row=row, column=decoding_col)
            cell.value = info["json_filename"]
            cell.hyperlink = link_path
            cell.font = LINK_FONT
        elif info["decode_error"]:
            cell = ws_files.cell(row=row, column=decoding_col)
            cell.value = info["decode_error"]
            cell.font = DEFAULT_FONT

        if info["type"] == "DF":
            for c in range(1, len(headers) + 1):
                ws_files.cell(row=row, column=c).fill = DF_FILL
        row += 1

    # LF/CF 레코드 시트 생성 + Bytes 셀에 링크 연결
    for file_path, (files_row, file_data) in record_sheets.items():
        path_parts = file_data.get("path", [])
        sheet_name = _create_record_sheet(wb, file_data, path_parts, files_row, used_sheet_names)

        # Bytes 셀을 시트 링크로 변경
        cell = ws_files.cell(row=files_row, column=bytes_col)
        cell.value = sheet_name
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = LINK_FONT

    # 컬럼 너비
    col_widths = [15, 15, 20, 8, 6, 8, 6, 10, 10, 50, 35, 25]
    for i, w in enumerate(col_widths):
        ws_files.column_dimensions[chr(65 + i)].width = w
    ws_files.auto_filter.ref = f"A1:{chr(64 + len(headers))}{row - 1}"

    # === Errors 시트 ===
    ws_errors = wb.create_sheet("Errors")
    _write_header(ws_errors, ["Path", "FID", "SW", "Message"])
    err_row = 2
    for file_path, file_data in _sorted_files(data):
        error = file_data.get("error")
        if error:
            fcp = file_data.get("fcp", {})
            vals = ["/".join(file_data.get("path", [])),
                    fcp.get("file_identifier", ""),
                    error.get("sw_actual", ""),
                    error.get("message", "")]
            for c, val in enumerate(vals, 1):
                cell = ws_errors.cell(row=err_row, column=c, value=val)
                cell.border = BORDER
                cell.fill = ERROR_FILL
            err_row += 1
    ws_errors.column_dimensions["A"].width = 40
    ws_errors.column_dimensions["B"].width = 8
    ws_errors.column_dimensions["C"].width = 10
    ws_errors.column_dimensions["D"].width = 40

    wb.save(str(output_path))
    logger.info("Excel 저장: %s (파일 %d개, 에러 %d개)", output_path, row - 2, err_row - 2)
    return str(output_path)


if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
    parser = argparse.ArgumentParser(description="SIM 덤프 JSON → Excel 변환")
    parser.add_argument("json_file", help="입력 JSON 파일 경로")
    parser.add_argument("-o", "--output", type=str, default=None)
    args = parser.parse_args()
    convert_to_excel(args.json_file, args.output)
